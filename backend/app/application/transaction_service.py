"""Service layer for Transaction operations."""

from __future__ import annotations

from decimal import Decimal
from uuid import UUID

from sqlalchemy.ext.asyncio import AsyncSession

from app.infrastructure.transaction_repository import TransactionRepository
from app.infrastructure.models import TransactionModel


class TransactionService:
    """Business logic for Transaction management."""

    def __init__(self, session: AsyncSession) -> None:
        self._repo = TransactionRepository(session)

    async def list_transactions(self, broker_id: UUID) -> list[TransactionModel]:
        """Return all transactions for a broker."""
        return await self._repo.get_by_broker(broker_id)

    async def create_transaction(self, data: dict) -> TransactionModel:
        """Create a new transaction with validation."""
        amount = data.get("amount")
        if amount is None or float(amount) <= 0:
            raise ValueError("Amount must be greater than 0")

        tx_type = data.get("type")
        valid_types = {"accrual", "payment", "transfer", "cash"}
        if tx_type not in valid_types:
            raise ValueError(f"Invalid transaction type: {tx_type}")

        return await self._repo.create(data)

    async def delete_transaction(self, transaction_id: UUID) -> bool:
        """Delete a transaction. Returns True if found and deleted."""
        return await self._repo.delete(transaction_id)

    async def get_debt(self, broker_id: UUID) -> Decimal:
        """
        Calculate current debt for a broker.
        Positive = debt, Negative = overpayment.
        """
        return await self._repo.calculate_debt(broker_id)

    async def create_many_transactions(self, data_list: list[dict]) -> list[TransactionModel]:
        """Create multiple transactions in bulk with validation."""
        for data in data_list:
            amount = data.get("amount")
            if amount is None or float(amount) <= 0:
                raise ValueError("Amount must be greater than 0")

            tx_type = data.get("type")
            valid_types = {"accrual", "payment", "transfer", "cash"}
            if tx_type not in valid_types:
                raise ValueError(f"Invalid transaction type: {tx_type}")

        return await self._repo.create_many(data_list)

    async def export_broker_transactions_to_excel(self, broker_id: UUID) -> bytes:
        """Export all transactions for a broker to an Excel file using openpyxl."""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        import io
        from datetime import timezone, timedelta

        # UTC+5 for reporting
        report_tz = timezone(timedelta(hours=5))

        transactions = await self.list_transactions(broker_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Транзакции"

        headers = [
            "Дата", "Тип", "Сумма", "Источник", 
            "Номер чека", "Отправитель", "Получатель", 
            "Комментарий", "КБК", "КНП", "Добавлен"
        ]
        
        ws.append(headers)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        type_map = {
            "accrual": "Начисление",
            "payment": "Оплата",
            "transfer": "Перевод",
            "cash": "Наличные (пополнение)"
        }

        source_map = {
            "manual": "Вручную",
            "receipt": "Чек (PDF)"
        }

        for tx in transactions:
            # Ensure timezone conversion for display
            display_dt = tx.datetime.astimezone(report_tz)
            display_created = tx.created_at.astimezone(report_tz)

            row = [
                display_dt.strftime("%d.%m.%Y %H:%M"),
                type_map.get(tx.type, tx.type),
                float(tx.amount),
                source_map.get(tx.source, tx.source),
                tx.receipt_number or "",
                tx.party_from or "",
                tx.party_to or "",
                tx.comment or "",
                tx.kbk or "",
                tx.knp or "",
                display_created.strftime("%d.%m.%Y %H:%M")
            ]
            ws.append(row)

        for col_idx in range(1, len(headers) + 1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50:
                adjusted_width = 50 
            ws.column_dimensions[column_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        wb.close()
        return output.getvalue()
